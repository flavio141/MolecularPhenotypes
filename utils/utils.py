import openpyxl


def save_percentage(perc):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "PDB wt"
    sheet["B1"] = "Percentage Position"

    row = 2
    for keys, subkeys in perc.items():
        sheet.cell(row=row, column=1, value=keys)
        sheet.cell(row=row, column=2, value=subkeys)
        row += 1

    wb.save("embedding/additional_features/percentage.xlsx")


def save_metrics(results):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "Epoch"
    sheet["B1"] = "Matthews Train"

    row = 2
    for keys, subkeys in results.items():
        sheet.cell(row=row, column=1, value=keys)
        sheet.cell(row=row, column=2, value=subkeys)
        row += 1

    wb.save("results/training_test.xlsx")
