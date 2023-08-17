import os
import pickle
import urllib
import openpyxl
import requests
import numpy as np
import pandas as pd
import requests as r

from tqdm import tqdm
from unirep_emb import UniRep_embedding
from sklearn.metrics.pairwise import cosine_similarity
from Bio.PDB import MMCIFParser, PDBIO, PDBParser, PDBList

data = pd.read_csv('dataset/SNV.tsv', sep='\t')

def save_percentage(perc):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet["A1"] = "PDB wt"
    sheet["B1"] = "Percentage Position"

    row = 2
    for keys, subkeys in perc.items():
        sheet.cell(row=row, column=1, value=keys)
        sheet.cell(row=row, column=2, value=subkeys)
        row += 1

    wb.save("embedding/additional_features/percentage.xlsx")


def download_cif_file(cIDs, pdbs):
    try:
        for cID, step in zip(cIDs, tqdm(range(0, len(cIDs)), desc= 'Extracting CIF and Converting')):
            pdb_id = cID.split(':')[0]
            cif_file = '{}.cif'.format(pdb_id)
            print(f'Download {cif_file}')

            if (pdb_id + '_' + cID.split(':')[1] + '.pdb') in os.listdir('dataset/pdb'):
                continue

            url = 'https://files.rcsb.org/download/{}.cif'.format(pdb_id)
            response = urllib.request.urlopen(url)
            cif_data = response.read().decode('utf-8')
            
            with open(os.path.join('dataset/cif',cif_file), 'w') as output_handle:
                output_handle.write(cif_data)
            
            chains = [pdb for pdb in pdbs if cID.split('.')[0] in pdb]

            for chain in chains:
                parser = MMCIFParser(QUIET=True)
                structure = parser.get_structure(pdb_id, os.path.join('dataset/cif', cif_file))

                selected_chain = None
                for element in structure[0]:
                    if element.get_id() == chain.split(':')[1]:
                        selected_chain = element
                        if len(selected_chain.get_id()) > 1:
                            selected_chain.id = '-'
                        break
                
                pdb_io = PDBIO()

                pdb_io.set_structure(selected_chain)
                pdb_save = f'dataset/pdb/{pdb_id}_{chain.split(":")[1]}.pdb'
                pdb_io.save(pdb_save)
    except Exception as error:
        print(f'There was an error: {error}')


def download_pdb_file(cIDs):
    try:
        not_pdb = []

        for cID, step in zip(cIDs, tqdm(range(0, len(cIDs)), desc= 'Extracting PDB using UniProt ID')):
            pdb_id = cID.split(':')[0]
            chain_id = cID.split(':')[1]
            pdb_list = PDBList(verbose=False)

            if (pdb_id + '_' + chain_id + '.pdb') in os.listdir('dataset/pdb'):
                continue
    
            pdb_url = f'https://files.rcsb.org/download/{pdb_id}.pdb'
            
            pdb_response = r.get(pdb_url)
            if pdb_response.status_code == 200:
                with open(os.path.join('dataset/pdb_temp', f'{pdb_id}.pdb'), 'wb') as file:
                    file.write(pdb_response.content)
            else:
                not_pdb.append(f'{pdb_id}:{chain_id}.pdb')
                continue

            pdb_list.retrieve_pdb_file(pdb_id, pdir="dataset/pdb_temp", file_format="pdb")
            parser = PDBParser(QUIET=True)

            structure = parser.get_structure(pdb_id, f"dataset/pdb_temp/{pdb_id}.pdb")

            selected_chain = None
            for chain in structure[0]:
                if chain.get_id() == chain_id:
                    selected_chain = chain
                    break

            io = PDBIO()
            io.set_structure(selected_chain)
            io.save(f"dataset/pdb/{pdb_id}_{chain_id}.pdb")

        return not_pdb
    except r.exceptions.RequestException as e:
        print(f'Error {cID}: {e}')


def download_uniprot_sequences():
    try:
        pdb_map = data.groupby('uniprot_id')['pdb_id'].unique()
        for uniprot_id, step in zip(data['uniprot_id'].unique(), tqdm(range(0, len(data['uniprot_id'].unique())), desc= 'Extracting FASTA')):
            url = f"https://www.uniprot.org/uniprot/{uniprot_id}.fasta"

            skip = False
            
            for pdb_name in pdb_map[uniprot_id]:
                if (pdb_name.replace(':', '_') + '.fasta') in os.listdir('dataset/fasta'):
                    skip = True
            
            if skip:
                continue

            response = requests.get(url)
            fasta_data = response.text.strip()

            fasta_data = fasta_data.replace('.', '')
            fasta_data = fasta_data.split('\n')

            for pdb in pdb_map[uniprot_id]:
                fasta = f'>{uniprot_id}\n{"".join(fasta_data[1:])}'
                output_file = f"dataset/fasta/{pdb.replace(':', '_')}.fasta"
                with open(output_file, 'w') as file:
                    file.write(fasta)
    except Exception as error:
        print(f'There was an error: {error}')


def fasta_mutated(wildtype, position, mutation, pdb):
    pdb_errors = open('pdb_errors.txt', 'w')

    for wt, pos, mut, pdb, count in zip(wildtype, position, mutation, pdb, tqdm(range(0, len(pdb)), desc= 'Extracting FASTA Mutated')):
        pdb_id = pdb.split(':')[0]
        chain = pdb.split(':')[1]

        with open(f'dataset/fasta/{pdb_id}_{chain}.fasta', 'r') as fasta:
            fasta_original = fasta.read()
        
        if len(fasta_original.split('\n')[1]) >= pos and fasta_original.split('\n')[1][pos - 1] == wt:
            if ',' in mut:
                for m in mut.split(','):
                    if f'{pdb_id}_{chain}_{wt}_{pos}_{m}.fasta' in os.listdir('dataset/fasta_mut'):
                        continue
                    mutation = m
                    fasta_seq = fasta_original.split('\n')[1][:pos - 1] + mutation + fasta_original.split('\n')[1][pos:]
                    fasta_mut = fasta_original.split('\n')[0] + '\n' + fasta_seq

                    with open(f'dataset/fasta_mut/{pdb_id}_{chain}_{wt}_{pos}_{m}.fasta', 'w') as mutated:
                        mutated.write(fasta_mut)
            else:
                mutation = mut
                if f'{pdb_id}_{chain}_{wt}_{pos}_{mut}.fasta' in os.listdir('dataset/fasta_mut'):
                        continue
                fasta_seq = fasta_original.split('\n')[1][:pos - 1] + mutation + fasta_original.split('\n')[1][pos:]
                fasta_mut = fasta_original.split('\n')[0] + '\n' + fasta_seq

                with open(f'dataset/fasta_mut/{pdb_id}_{chain}_{wt}_{pos}_{mut}.fasta', 'w') as mutated:
                    mutated.write(fasta_mut)
        elif len(fasta_original.split('\n')[1]) >= pos:
            original = fasta_original.split("\n")[1][pos - 1]
            pdb_errors.write(f'{pdb} in position {pos}')
            print(f'No match between {wt} and the amino acids at position {original}:{pos} for {pdb}')
        else:
            pdb_errors.write(f'{pdb} in position {pos}')
            print(f'The position is out of range: {pos} and the pdb is {pdb}')


def feature_extraction_mut(cIDs):
    for cID, step in zip(cIDs, tqdm(range(0, len(cIDs)), desc= 'Extracting Features for FASTA Mutated')):
        pdb = cID.split('.')[0]
        path_fasta = f'dataset/fasta_mut/{pdb}.fasta'
        path_fasta_emb = f'embedding/fastaEmb_mut/{pdb}.embeddings.pkl'
        if not f'{pdb}.embeddings.pkl' in os.listdir('embedding/fastaEmb_mut'):
            os.system(f'python GCN-for-Structure-and-Function/scripts/seqvec_embedder.py --input={path_fasta} --output={path_fasta_emb}')


def feature_extraction_wt(cIDs):
    try:
        for cID, step in zip(cIDs, tqdm(range(0, len(cIDs)), desc= 'Extracting Features for WildType')):
            pdb = '_'.join(cID.split(':'))
            path_fasta = f'dataset/fasta/{pdb}.fasta'
            path_fasta_emb = f'embedding/fastaEmb_wt/{pdb}.embeddings.pkl'
            if not f'{pdb}.embeddings.pkl' in os.listdir('embedding/fastaEmb_wt'):
                os.system(f'python GCN-for-Structure-and-Function/scripts/seqvec_embedder.py --input={path_fasta} --output={path_fasta_emb}')

            path_pdb = f'dataset/pdb/{pdb}.pdb'
            path_pdb_emb = f'embedding/distmap_wt/{pdb}.distmap.npy'
            if not f'{pdb}.distmap.npy' in os.listdir('embedding/distmap_wt'):
                os.system(f'python GCN-for-Structure-and-Function/scripts/convert_pdb_to_distmap.py {path_pdb} {path_pdb_emb}')
    except Exception as error:
        print(f'There was an error: {error}')


def similarity():
    cos_similarity = {}

    try:
        for pdb, wt, pos, mut, count in zip(data['pdb_id'], data['wildtype'], data['position'], data['mutation'], tqdm(range(0, len(data['mutation'])), desc= 'Similarity')):
            pdb_id = "_".join(pdb.split(':'))

            if pdb.split(':')[0] == '4jnw': # perché ha 32000 amminoacidi e ci vuole troppo, non riesce a gestirlo
                continue
            
            if pdb_id not in cos_similarity.keys():
                cos_similarity[pdb_id] = {}

            if f'{pdb_id}.embeddings.pkl' not in os.listdir('embedding/fastaEmb_wt'):
                continue

            with open (f'embedding/fastaEmb_wt/{pdb_id}.embeddings.pkl', 'rb') as wild:
                wildtype = pickle.load(wild)


            if ',' in mut:
                for m in mut.split(','):
                    
                    if f'{pdb_id}_{wt}_{pos}_{m}.embeddings.pkl' not in os.listdir('embedding/fastaEmb_mut'):
                        continue

                    with open (f'embedding/fastaEmb_mut/{pdb_id}_{wt}_{pos}_{m}.embeddings.pkl', 'rb') as mt:
                        mutated = pickle.load(mt)
                    similarity_matrix = cosine_similarity(wildtype[list(wildtype.keys())[0]], mutated[list(mutated.keys())[0]])
                    cos_similarity[pdb_id][pdb_id + '_' + wt + '_' + str(pos) + '_' + m] = [similarity_matrix]

                    cos_similarity[pdb_id][pdb_id + '_' + wt + '_' + str(pos) + '_' + m].append(wildtype[list(wildtype.keys())[0]] - mutated[list(mutated.keys())[0]])
            else:
                if f'{pdb_id}_{wt}_{pos}_{mut}.embeddings.pkl' not in os.listdir('embedding/fastaEmb_mut'):
                    continue

                with open (f'embedding/fastaEmb_mut/{pdb_id}_{wt}_{pos}_{mut}.embeddings.pkl', 'rb') as mt:
                    mutated = pickle.load(mt)
                
                similarity_matrix = cosine_similarity(wildtype[list(wildtype.keys())[0]], mutated[list(mutated.keys())[0]])
                cos_similarity[pdb_id][pdb_id + '_' + wt + '_' + str(pos) + '_' + mut] = [similarity_matrix]

                cos_similarity[pdb_id][pdb_id + '_' + wt + '_' + str(pos) + '_' + mut].append(wildtype[list(wildtype.keys())[0]] - mutated[list(mutated.keys())[0]])
        
        np.save('embedding/additional_features/similarity.npy',cos_similarity)
    except Exception as error:
        print(f'There was an error: {error} {pdb_id}')


def percentage():
    percentage_dict = {}
    for fasta in os.listdir('dataset/fasta_mut'):
        sequence = ''
        with open(f'dataset/fasta_mut/{fasta}', 'r') as file:
            for line in file:
                line = line.strip()
                if line.startswith('>'):
                    continue
                sequence += line
        position = int(fasta.split('_')[3])
        perc = position / len(sequence) * 100

        percentage_dict[fasta.split('.')[0]] = perc
    
    save_percentage(percentage_dict)


def extract_data_wt(extract_data=False):
    if not os.path.exists('dataset'):
        raise
    
    if extract_data:
        try:
            #data.drop(columns=['phenotypic_annotation'], inplace=True)

            nans = {'wildtype' : data['wildtype'].isnull().sum(), 
                    'position': data['position'].isnull().sum(), 
                    'mutation': data['mutation'].isnull().sum()
                }
            
            if nans['wildtype'] == 0 or nans['position'] == 0 or nans['mutation'] == 0:
                data.dropna(subset = [min(nans, key=nans.get)])

            # Download all PDB files
            not_pdb = download_pdb_file(data['pdb_id'].unique())

            # Download PDB files too large as CIF and then convert them
            if not_pdb:
                download_cif_file(list(set(not_pdb)), data['pdb_id'].unique())
            
            # Extract FASTA
            download_uniprot_sequences()

            # Extract Features
            feature_extraction_wt(data['pdb_id'].unique())
        except Exception as error:
            print(f'There was an error: {error}')
    else:
        print('Data not extracted as declared')


def extract_data_mut(extract_data_mut=False):
    if extract_data_mut:
        try:
            fasta_mutated(data['wildtype'], data['position'], data['mutation'], data['pdb_id'])
            feature_extraction_mut(os.listdir('dataset/fasta_mut'))
        except Exception as error:
            print(f'There was an error: {error}')
    else:
        print('Data not extracted as declared')


def features(args, folders):
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(folder)

    extract_data_wt(args.extract_data_wt)
    extract_data_mut(args.extract_data_mut)

    if args.extract_unirep == True:
        UniRep_embedding(os.listdir('dataset/fasta'), os.listdir('dataset/fasta_mut'))

    if args.extract_similarity == True:
        similarity()
    if args.extract_percentage == True:
        percentage()
    